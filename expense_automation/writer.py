"""Excel rendering utilities for claim forms."""

from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import List

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from .models import ClaimHeader, ExpenseItem


class ClaimFormWriter:
    """Write expenses into a paginated Excel claim form template."""

    MAX_ROWS = 10
    START_ROW = 13
    END_ROW = 22

    COL_MAP = {
        "sn": 2,
        "transaction_date": 3,
        "supplier_name": 4,
        "invoice_no": 5,
        "expense_type": 6,
        "description": 7,
        "currency": 8,
        "original_amount": 9,
        "forex_rate": 10,
        "amount_before_gst_sgd": 11,
        "gst_amount": 12,
        "final_claim_amount_sgd": 13,
    }

    def __init__(self, template_path: str, output_path: str):
        self.template_path = Path(template_path)
        self.output_path = Path(output_path)
        self.workbook = None

    def load_template(self) -> None:
        if not self.template_path.exists():
            raise FileNotFoundError(
                f"Template file '{self.template_path}' is missing. Please provide the blank claim form."
            )
        self.workbook = openpyxl.load_workbook(self.template_path)

    def _write_header_footer(self, worksheet: Worksheet, header: ClaimHeader) -> None:
        # 静态头尾字段填充，对应 HR 模板固定位置
        worksheet["M2"] = header.month_of_claim
        worksheet["C29"] = header.employee_name
        worksheet["C30"] = header.department
        worksheet["I29"] = header.approver_name
        worksheet["C32"] = header.signature_date()

    def _clear_row(self, worksheet: Worksheet, row_idx: int) -> None:
        # 清空数据但保留原有单元格格式
        for col in range(self.COL_MAP["sn"], self.COL_MAP["final_claim_amount_sgd"] + 1):
            worksheet.cell(row=row_idx, column=col, value=None)

    def _write_batch(self, worksheet: Worksheet, header: ClaimHeader, batch: List[ExpenseItem], offset: int) -> None:
        self._write_header_footer(worksheet, header)

        for index, item in enumerate(batch):
            row_number = self.START_ROW + index
            worksheet.cell(row=row_number, column=self.COL_MAP["sn"], value=offset + index + 1)
            worksheet.cell(row=row_number, column=self.COL_MAP["transaction_date"], value=item.transaction_date)
            worksheet.cell(row=row_number, column=self.COL_MAP["supplier_name"], value=item.supplier_name)
            worksheet.cell(row=row_number, column=self.COL_MAP["invoice_no"], value=item.invoice_no)
            worksheet.cell(row=row_number, column=self.COL_MAP["expense_type"], value=item.expense_type)
            worksheet.cell(row=row_number, column=self.COL_MAP["description"], value=item.description)
            worksheet.cell(row=row_number, column=self.COL_MAP["currency"], value=item.currency)
            worksheet.cell(row=row_number, column=self.COL_MAP["original_amount"], value=float(item.original_amount))
            worksheet.cell(row=row_number, column=self.COL_MAP["forex_rate"], value=float(item.forex_rate))
            worksheet.cell(
                row=row_number,
                column=self.COL_MAP["amount_before_gst_sgd"],
                value=float(item.amount_before_gst_sgd or 0),
            )
            worksheet.cell(row=row_number, column=self.COL_MAP["gst_amount"], value=float(item.gst_amount))
            worksheet.cell(
                row=row_number,
                column=self.COL_MAP["final_claim_amount_sgd"],
                value=float(item.final_claim_amount_sgd or 0),
            )

        for row_idx in range(self.START_ROW + len(batch), self.END_ROW + 1):
            self._clear_row(worksheet, row_idx)

    def process(self, header: ClaimHeader, items: List[ExpenseItem]) -> None:
        self.load_template()

        sorted_items = sorted(items, key=lambda item: item.transaction_date)
        total_items = len(sorted_items)
        num_batches = max(1, (total_items // self.MAX_ROWS) + (1 if total_items % self.MAX_ROWS else 0))

        base_sheet = self.workbook.active
        base_sheet.title = "Page_1"

        for batch_index in range(num_batches):
            if batch_index == 0:
                sheet = base_sheet
            else:
                sheet = self.workbook.copy_worksheet(base_sheet)
                sheet.title = f"Page_{batch_index + 1}"

            start_index = batch_index * self.MAX_ROWS
            end_index = start_index + self.MAX_ROWS
            batch = sorted_items[start_index:end_index]
            self._write_batch(sheet, header, batch, offset=start_index)

        self.workbook.save(self.output_path)


__all__ = ["ClaimFormWriter"]

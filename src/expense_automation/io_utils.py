import csv
import json
from datetime import date, datetime
from pathlib import Path
from typing import Iterable, List, Dict, Any, Tuple

DateKey = Tuple[str, str]

REQUIRED_RECEIPT_FIELDS = {"receipt_id", "date", "plate", "merchant", "amount", "category"}
REQUIRED_EXTERNAL_FIELDS = {"plate", "date", "source", "amount", "note"}


def _parse_date(value: str | datetime | date) -> str:
    """Normalize date-like values to a YYYY-MM-DD string."""

    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    try:
        return datetime.strptime(str(value).strip(), "%Y-%m-%d").date().isoformat()
    except ValueError as exc:
        raise ValueError(f"Invalid date '{value}'. Expected YYYY-MM-DD.") from exc


def load_receipts(path: Path) -> List[Dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".json":
        return _load_receipts_json(path)
    if suffix == ".csv":
        return _load_receipts_csv(path)
    if suffix in {".xlsx", ".xls"}:
        return _load_receipts_excel(path)
    if suffix in {".png", ".jpg", ".jpeg", ".webp"}:
        return _load_receipts_from_images([path])
    raise ValueError(
        f"Unsupported receipt file type: {suffix}. Use JSON, CSV, Excel, or image files."
    )


def _load_receipts_json(path: Path) -> List[Dict[str, Any]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    return _normalize_receipts(data)


def _load_receipts_csv(path: Path) -> List[Dict[str, Any]]:
    with path.open("r", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames is None or not REQUIRED_RECEIPT_FIELDS.issubset(reader.fieldnames):
            raise ValueError(
                f"Receipt CSV must contain columns: {sorted(REQUIRED_RECEIPT_FIELDS)}"
            )
        return _normalize_receipts(list(reader))


def _load_receipts_excel(path: Path) -> List[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # noqa: BLE001
        raise ImportError(
            "Reading Excel files requires the optional dependency 'openpyxl'."
        ) from exc

    wb = load_workbook(filename=path, read_only=True, data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(col).strip() for col in rows[0]]
    if not REQUIRED_RECEIPT_FIELDS.issubset(headers):
        raise ValueError(
            f"Receipt Excel must contain columns: {sorted(REQUIRED_RECEIPT_FIELDS)}"
        )
    records = [dict(zip(headers, row)) for row in rows[1:]]
    return _normalize_receipts(records)


def _load_receipts_from_images(paths: List[Path]) -> List[Dict[str, Any]]:
    today = date.today().isoformat()
    receipts: List[Dict[str, Any]] = []
    for idx, path in enumerate(paths, start=1):
        receipts.append(
            {
                "receipt_id": path.stem or f"image-{idx}",
                "date": today,
                "plate": "UNKNOWN",
                "merchant": "image_upload",
                "amount": 0.0,
                "category": "image_placeholder",
            }
        )
    return receipts


def _normalize_receipts(data: Iterable[Dict[str, Any]]) -> List[Dict[str, Any]]:
    if not isinstance(data, Iterable):
        raise ValueError("Receipt payload must be iterable.")

    receipts: List[Dict[str, Any]] = []
    for idx, raw in enumerate(data, start=1):
        if not REQUIRED_RECEIPT_FIELDS.issubset(raw):
            missing = REQUIRED_RECEIPT_FIELDS - set(raw)
            raise ValueError(f"Receipt #{idx} is missing fields: {sorted(missing)}")

        normalized = {
            "receipt_id": str(raw["receipt_id"]).strip(),
            "date": _parse_date(raw["date"]),
            "plate": str(raw["plate"]).strip(),
            "merchant": str(raw["merchant"]).strip(),
            "amount": float(raw["amount"]),
            "category": str(raw["category"]).strip(),
        }
        receipts.append(normalized)
    return receipts


def load_external_charges(path: Path) -> Dict[DateKey, List[Dict[str, Any]]]:
    if not path.exists():
        raise FileNotFoundError(f"External data file not found: {path}")

    suffix = path.suffix.lower()
    if suffix == ".csv":
        records = _load_external_csv(path)
    elif suffix in {".xlsx", ".xls"}:
        records = _load_external_excel(path)
    else:
        raise ValueError(
            f"Unsupported external data type: {suffix}. Use CSV or Excel files."
        )

    charges: Dict[DateKey, List[Dict[str, Any]]] = {}
    for idx, record in enumerate(records, start=1):
        try:
            normalized_date = _parse_date(record["date"])
            normalized = {
                "plate": str(record["plate"]).strip(),
                "date": normalized_date,
                "source": str(record["source"]).strip(),
                "amount": float(record["amount"]),
                "note": str(record.get("note", "")).strip(),
            }
        except Exception as exc:  # noqa: BLE001
            raise ValueError(f"Invalid external record at index {idx}: {record}") from exc

        key: DateKey = (normalized["plate"], normalized["date"])
        charges.setdefault(key, []).append(normalized)
    return charges


def _load_external_csv(path: Path) -> List[Dict[str, Any]]:
    with path.open("r", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames is None or not REQUIRED_EXTERNAL_FIELDS.issubset(reader.fieldnames):
            raise ValueError(
                f"CSV must contain columns: {sorted(REQUIRED_EXTERNAL_FIELDS)}"
            )
        return list(reader)


def _load_external_excel(path: Path) -> List[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # noqa: BLE001
        raise ImportError(
            "Reading Excel files requires the optional dependency 'openpyxl'."
        ) from exc

    wb = load_workbook(filename=path, read_only=True, data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(col).strip() for col in rows[0]]
    if not REQUIRED_EXTERNAL_FIELDS.issubset(headers):
        raise ValueError(
            f"External Excel must contain columns: {sorted(REQUIRED_EXTERNAL_FIELDS)}"
        )
    return [dict(zip(headers, row)) for row in rows[1:]]


def ensure_output_directory(path: Path) -> None:
    target_dir = path.parent
    target_dir.mkdir(parents=True, exist_ok=True)


def export_to_csv(records: Iterable[Dict[str, Any]], output_path: Path) -> None:
    ensure_output_directory(output_path)
    records = list(records)
    if not records:
        output_path.write_text("", encoding="utf-8")
        return

    fieldnames = list(records[0].keys())
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(records)
